import openpyxl


def load_excel(filepath, sheet):
    global wb, sh
    wb = openpyxl.load_workbook(filepath)
    sh = wb[sheet]


def get_cell_data(rowno, colno):
    return sh.cell(rowno, colno).value


def get_data_as_list_tuple():
    sheet_cells = []  # list of tuple to hold multiple rows data
    for i in range(1, sh.max_row):
        row_cells = []
        for cell in sh[i+1]:
            row_cells.append(cell.value)
        sheet_cells.append(tuple(row_cells))
    return sheet_cells
