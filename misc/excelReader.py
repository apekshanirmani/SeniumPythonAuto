import openpyxl

wb = openpyxl.load_workbook("C:/SWEDEN/Nirmani/Automation/PythonProject/SeniumPythonAuto/test_data.xlsx")
sh = wb['Sheet1']

""""
# print single cell value
print(sh.cell(1, 1).value)

# print multiple cell values
for cell in sh['A']:
    print(cell.value)

# print values value in the loop
for i in range(1, sh.max_row+1):
    for cell in sh[i]:
        print(cell.value)

for row in sh.iter_rows(1,sh.max_row):
    for cell in row:
        print(cell.value)

sheet_cells = [] # list of tuple to hold multiple rows data
for row in sh.iter_rows():
    row_cells = []
    for cells in row:
        row_cells.append(cells.value)
    sheet_cells.append(tuple(row_cells))
print(sheet_cells)

for i in range(1, len(sheet_cells)):
    print(sheet_cells[i])
"""

# create new sheet and write some data
wb.create_sheet('data')
sh1 = wb['data']

sh1.cell(1, 1).value = "Nir"
print(sh1.cell(1, 1).value)

wb.save("C:/SWEDEN/Nirmani/Automation/PythonProject/SeniumPythonAuto/test_data.xlsx")
